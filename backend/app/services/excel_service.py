from __future__ import annotations

import re
import zipfile
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import Outline

from app.models import APPROVED_ITEM_STATUSES, EXPORTABLE_REQUEST_STATUSES
from app.repositories.base import Repository
from app.services.common import get_required, require_role
from app.services.file_service import FileService
from app.services.file_guard_client import FileGuardClient, require_valid_file
from app.services.permission_service import PermissionService
from app.services.request_service import RequestService


HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DEPARTMENT_FILL = PatternFill("solid", fgColor="EAF1FF")
CFO_FILL = PatternFill("solid", fgColor="F5F8FE")
ARTICLE_FILL = PatternFill("solid", fgColor="FAFBFD")
MONEY_FORMAT = '#,##0.00'

REQUEST_STATUS_LABELS = {
    "draft": "Черновик",
    "on_review": "На проверке",
    "approved": "Утверждена",
    "rejected": "Отклонена",
    "cancelled": "Отменена",
}

ITEM_STATUS_LABELS = {
    "on_review": "На рассмотрении",
    "rejected": "Отказано",
    "approved_with_changes": "Утверждено с изменениями",
    "approved": "Утверждено",
}

ANALYTICS_FIELDS = tuple(f"analytics_{index}" for index in range(1, 6))
ANALYTICS_HEADERS = [f"Аналитика {index}" for index in range(1, 6)]
REGISTER_VIEW_LABELS = {
    "cfo": "По ЦФО",
    "category": "По категории",
    "article": "По статье",
    "module": "По модулю",
    "request": "По заявкам",
}
REGISTER_GROUP_LABELS = {
    "cfo": "ЦФО",
    "article": "Статья",
    "category": "Категория",
    "module": "Модуль",
    "request": "Заявка",
    **{field: f"Аналитика {field[-1]}" for field in ANALYTICS_FIELDS},
}
AGGREGATE_STATUS_LABELS = {
    "approved": "Согласовано",
    "rejected": "Отклонено",
    "partially_approved": "Частично рассмотрено",
    "on_review": "На рассмотрении",
    "in_progress": "Частично рассмотрено",
    "no_data": "Черновик",
}
SUMMARY_FILL = PatternFill("solid", fgColor="F8FAFC")
NOTE_FONT = Font(italic=True, color="64748B")
GROUP_FILLS = (DEPARTMENT_FILL, CFO_FILL, ARTICLE_FILL)
MONTH_HEADERS = (
    "Январь",
    "Февраль",
    "Март",
    "Апрель",
    "Май",
    "Июнь",
    "Июль",
    "Август",
    "Сентябрь",
    "Октябрь",
    "Ноябрь",
    "Декабрь",
)


class ExcelService:
    def __init__(
        self,
        repo: Repository,
        permissions: PermissionService,
        requests: RequestService,
        files: FileService,
        export_dir: Path,
        file_guard: FileGuardClient,
    ):
        self.repo = repo
        self.permissions = permissions
        self.requests = requests
        self.files = files
        self.export_dir = export_dir
        self.file_guard = file_guard
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def department_id_for_unit(self, unit_id: str | None) -> str | None:
        if not unit_id:
            return None
        unit = self.repo.get_by_id("units", unit_id)
        if not unit:
            return None
        if unit.get("parent_id"):
            return unit["parent_id"]
        return unit["id"]

    def resolve_unit_id(self, *, unit_id: str | None = None, module_id: str | None = None) -> str | None:
        if module_id:
            return self.department_id_for_unit(module_id)
        if unit_id:
            return self.department_id_for_unit(unit_id)
        return None

    def filter_catalog(
        self,
        collection: str,
        *,
        unit_id: str | None = None,
        module_id: str | None = None,
        active_only: bool = False,
        query: str | None = None,
    ) -> list[dict]:
        department_id = self.resolve_unit_id(unit_id=unit_id, module_id=module_id)
        items = self.repo.load_all(collection)
        result = []
        needle = (query or "").strip().lower()
        for item in items:
            if department_id and item.get("unit_id") not in {department_id, None}:
                # Allow global items (unit_id null) for admins browsing everything;
                # for scoped lookups require matching department.
                if unit_id or module_id:
                    if item.get("unit_id") != department_id:
                        continue
            if active_only and not item.get("is_active", True):
                continue
            if needle:
                haystack = str(item.get("name", "")).lower()
                if needle not in haystack:
                    continue
            result.append(item)
        return result

    @staticmethod
    def _style_header(ws, columns: list[str]) -> None:
        for index, title in enumerate(columns, start=1):
            cell = ws.cell(1, index, title)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _autosize(ws) -> None:
        for column in ws.columns:
            width = 12
            letter = column[0].column_letter
            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(value) + 2, 48))
            ws.column_dimensions[letter].width = width

    @staticmethod
    def _even_month_amounts(annual: float) -> list[float]:
        total_cents = int(Decimal(str(annual or 0)) * 100)
        base, remainder = divmod(total_cents, 12)
        return [(base + (1 if month <= remainder else 0)) / 100 for month in range(1, 13)]

    @staticmethod
    def _empty_months() -> list[float]:
        return [0.0] * 12

    @staticmethod
    def _add_months(left: list[float], right: list[float]) -> list[float]:
        return [left[index] + right[index] for index in range(12)]

    def _month_amounts_by_item(
        self,
        item_ids: set[str],
        annual_by_id: dict[str, float],
    ) -> dict[str, list[float]]:
        stored: dict[str, dict[int, float]] = {}
        for row in self.repo.load_all("req_item_month_plans"):
            item_id = row.get("req_item_id")
            if item_id not in item_ids:
                continue
            stored.setdefault(item_id, {})[int(row["month"])] = float(row.get("sum_plan") or 0)
        amounts: dict[str, list[float]] = {}
        for item_id in item_ids:
            by_month = stored.get(item_id)
            if by_month:
                amounts[item_id] = [float(by_month.get(month, 0)) for month in range(1, 13)]
            else:
                amounts[item_id] = self._even_month_amounts(annual_by_id.get(item_id, 0))
        return amounts

    @staticmethod
    def _hide_columns(ws, start: int, count: int) -> None:
        for index in range(start, start + count):
            ws.column_dimensions[get_column_letter(index)].hidden = True

    @staticmethod
    def _format_money_columns(ws, columns: range | tuple[int, ...] | list[int]) -> None:
        for column in columns:
            for row in range(2, (ws.max_row or 1) + 1):
                ws.cell(row, column).number_format = MONEY_FORMAT

    def build_import_template(self, kind: str) -> BytesIO:
        titles = {
            "dds": "Шаблон импорта статей ДДС",
            "invests": "Шаблон импорта инвест-проектов",
        }
        if kind not in titles:
            raise HTTPException(status_code=400, detail="Неизвестный тип справочника")
        leaf_label = "Статья ДДС" if kind == "dds" else "Инвест-проект"
        wb = Workbook()
        ws = wb.active
        ws.title = "НСИ"
        columns = ["Категория", "Название", "Подразделение", "Активен"]
        self._style_header(ws, columns)
        ws.append(["Операционные расходы", f"Пример: {leaf_label}", "Департамент цифровых продуктов", "да"])
        ws.append(["Операционные расходы", "Ещё одна подкатегория", "Департамент цифровых продуктов", "да"])
        ws.append(["Капитальные затраты", "Подкатегория другой категории", "Департамент цифровых продуктов", "да"])
        note = wb.create_sheet("Инструкция")
        note["A1"] = titles[kind]
        note["A2"] = "Структура НСИ: категория → подкатегория (статья ДДС или инвест-проект)."
        note["A3"] = "Обязательные поля: Категория, Название. Рекомендуется Подразделение."
        note["A4"] = "Одинаковая Категория в нескольких строках создаёт одну категорию и несколько подкатегорий."
        note["A5"] = "Подразделение должно совпадать с названием подразделения (корневого unit)."
        note["A6"] = "Активен: да/нет, true/false, 1/0."
        ws.delete_rows(1, ws.max_row)
        article_label = "Статья ДДС" if kind == "dds" else "Инвест-проект"
        self._style_header(ws, [article_label, "Категория", "Объединение", "Активен"])
        ws.append(["Операционные расходы", "Операционные расходы", "", "да"])
        ws.append(["Операционные расходы", "Подписки", "", "да"])
        note["A1"] = "Структура НСИ: статья / инвест-проект → категория."
        note["A2"] = "Если категория не указана, одноимённая категория будет создана только у статьи без других категорий."
        note["A3"] = "Повторяющаяся статья создаёт дополнительные категории под той же статьёй."
        note["A4"] = None
        note["A5"] = None
        note["A6"] = None
        self._autosize(ws)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _find_unit(self, unit_name: str | None, unit_id: str | None) -> str | None:
        if unit_id:
            unit = self.repo.get_by_id("units", str(unit_id))
            if not unit:
                raise HTTPException(status_code=400, detail=f"Подразделение {unit_id} не найдено")
            return unit["id"] if not unit.get("parent_id") else unit["parent_id"]
        if unit_name:
            name = str(unit_name).strip().lower()
            units = self.repo.load_all("units")
            match = next((unit for unit in units if not unit.get("parent_id") and unit.get("name", "").strip().lower() == name), None)
            if not match:
                match = next((unit for unit in units if unit.get("name", "").strip().lower() == name), None)
            if not match:
                raise HTTPException(status_code=400, detail=f"Подразделение «{unit_name}» не найдено")
            return match["id"] if not match.get("parent_id") else match["parent_id"]
        department = next((unit for unit in self.repo.load_all("units") if unit.get("parent_id") is None), None)
        return department["id"] if department else None

    @staticmethod
    def _as_bool(value: Any, default: bool = True) -> bool:
        if value is None or value == "":
            return default
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        if text in {"1", "true", "yes", "y", "да", "истина", "активен"}:
            return True
        if text in {"0", "false", "no", "n", "нет", "ложь", "неактивен"}:
            return False
        return default

    @staticmethod
    def _normalize_header(value: Any) -> str:
        text = str(value or "").strip().lower().replace(" ", "_")
        if text in {"статья_ддс", "инвест-проект", "статья_/_инвест-проект"}:
            return "name"
        mapping = {
            "название": "name",
            "наименование": "name",
            "имя": "name",
            "подкатегория": "name",
            "статья": "name",
            "проект": "name",
            "категория": "category",
            "category": "category",
            "category_name": "category",
            "родитель": "category",
            "parent": "category",
            "подразделение": "unit_name",
            "департамент": "unit_name",
            "unit": "unit_name",
            "unit_id": "unit_id",
            "активен": "is_active",
            "is_active": "is_active",
            "name": "name",
            "unit_name": "unit_name",
        }
        return mapping.get(text, text)

    def _ensure_article(
        self,
        collection: str,
        *,
        article_name: str,
        unit_id: str | None,
        is_active: bool,
    ) -> dict:
        name_key = article_name.strip()
        match = next(
            (
                item
                for item in self.repo.load_all(collection)
                if not item.get("parent_id")
                and item.get("name", "").strip().lower() == name_key.lower()
                and item.get("unit_id") == unit_id
            ),
            None,
        )
        if match:
            return match
        return self.repo.create(
            collection,
            {
                "parent_id": None,
                "name": name_key,
                "unit_id": unit_id,
                "is_active": is_active,
            },
        )

    def _find_leaf(self, collection: str, *, name: str, parent_id: str | None, unit_id: str | None) -> dict | None:
        return next(
            (
                item
                for item in self.repo.load_all(collection)
                if item.get("name", "").strip().lower() == name.strip().lower()
                and item.get("parent_id") == parent_id
                and item.get("unit_id") == unit_id
            ),
            None,
        )

    async def import_catalog(self, user: dict, collection: str, upload: UploadFile, *, preview: bool = False) -> dict:
        require_role(user, "admin")
        filename = (upload.filename or "").lower()
        if not filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Ожидается файл Excel (.xlsx)")
        await require_valid_file(self.file_guard, upload)
        raw = await upload.read()
        if not raw:
            raise HTTPException(status_code=400, detail="Пустой файл")
        try:
            wb = load_workbook(BytesIO(raw), data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=400, detail="Не удалось прочитать Excel-файл") from exc
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="В файле нет строк")
        headers = [self._normalize_header(value) for value in rows[0]]
        if "name" not in headers:
            raise HTTPException(status_code=400, detail="В первой строке должен быть столбец «Наименование», «Название» или «Подкатегория»")

        prepared: list[dict] = []
        errors: list[str] = []

        for row_index, values in enumerate(rows[1:], start=2):
            if not values or all(cell is None or str(cell).strip() == "" for cell in values):
                continue
            row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
            name = str(row.get("name") or "").strip()
            if not name:
                errors.append(f"Строка {row_index}: пустое название подкатегории")
                continue
            category_name = str(row.get("category") or "").strip() or None
            try:
                unit_id = self._find_unit(
                    str(row.get("unit_name")).strip() if row.get("unit_name") not in (None, "") else None,
                    str(row.get("unit_id")).strip() if row.get("unit_id") not in (None, "") else None,
                )
            except HTTPException as exc:
                errors.append(f"Строка {row_index}: {exc.detail}")
                continue

            prepared.append(
                {
                    "row": row_index,
                    "name": name,
                    "category": category_name,
                    "unit_id": unit_id,
                    "unit_name": str(row.get("unit_name") or "").strip(),
                    "is_active": self._as_bool(row.get("is_active"), True),
                }
            )

        # Импорт применяется только целиком: ошибки в любой строке не должны оставлять
        # в справочнике частично загруженные данные.
        if errors:
            return {
                "preview": preview,
                "created": 0,
                "updated": 0,
                "errors": errors,
                "rows": prepared,
                "collection": collection,
            }

        if preview:
            preview_rows = []
            created = 0
            updated = 0
            catalog = self.repo.load_all(collection)
            for item in prepared:
                parent = next(
                    (
                        entry
                        for entry in catalog
                if not entry.get("parent_id")
                        and entry.get("unit_id") == item["unit_id"]
                        and entry.get("name", "").strip().casefold() == item["name"].casefold()
                    ),
                    None,
                )
                category_name = item["category"]
                children = [entry for entry in catalog if parent and entry.get("parent_id") == parent["id"]]
                # A blank category asks for a fallback only while the article
                # has no children. This matches manual creation exactly.
                if not category_name and children:
                    action = "skip"
                else:
                    existing = self._find_leaf(
                        collection,
                        name=category_name or item["name"],
                        parent_id=parent["id"] if parent else None,
                        unit_id=item["unit_id"],
                    )
                    action = "update" if existing else "create"
                    updated += int(bool(existing))
                    created += int(not existing)
                preview_rows.append({**item, "action": action})
            return {
                "preview": True,
                "created": created,
                "updated": updated,
                "errors": [],
                "rows": preview_rows,
                "collection": collection,
            }

        created = 0
        updated = 0
        for item in prepared:
            parent = self._ensure_article(
                collection,
                article_name=item["name"],
                unit_id=item["unit_id"],
                is_active=True,
            )
            children = [
                entry for entry in self.repo.load_all(collection)
                if entry.get("parent_id") == parent["id"]
            ]
            # Do not append an article-name fallback after any category exists.
            if not item["category"] and children:
                continue
            payload = {
                "name": item["category"] or item["name"],
                "parent_id": parent["id"],
                "unit_id": item["unit_id"],
                "is_active": item["is_active"],
            }
            existing = self._find_leaf(
                collection,
                name=payload["name"],
                parent_id=payload["parent_id"],
                unit_id=item["unit_id"],
            )
            if existing:
                self.repo.update(collection, existing["id"], payload)
                updated += 1
            else:
                self.repo.create(collection, payload)
                created += 1

        return {
            "preview": False,
            "created": created,
            "updated": updated,
            "errors": [],
            "rows": prepared,
            "collection": collection,
        }

    def _unit_name(self, unit_id: str | None) -> str:
        if not unit_id:
            return ""
        unit = self.repo.get_by_id("units", unit_id)
        return unit.get("name", unit_id) if unit else unit_id

    def _department_name(self, unit_id: str | None) -> str:
        current_id = unit_id
        visited: set[str] = set()
        while current_id and current_id not in visited:
            visited.add(current_id)
            unit = self.repo.get_by_id("units", current_id)
            if not unit:
                return current_id
            if not unit.get("parent_id"):
                return unit.get("name", current_id)
            current_id = unit["parent_id"]
        return ""

    def _unit_hierarchy(self, unit_id: str | None) -> list[dict]:
        hierarchy: list[dict] = []
        current_id = unit_id
        visited: set[str] = set()
        while current_id and current_id not in visited:
            visited.add(current_id)
            unit = self.repo.get_by_id("units", current_id)
            if not unit:
                break
            hierarchy.append(unit)
            current_id = unit.get("parent_id")
        return list(reversed(hierarchy))

    def _zgd_unit_groups(self, unit_id: str | None) -> tuple[str, str, str]:
        hierarchy = self._unit_hierarchy(unit_id)
        if not hierarchy:
            return "Не указано", "Не указано", self._unit_name(unit_id)
        department = hierarchy[0].get("name") or "Не указано"
        cfo = (hierarchy[-2] if len(hierarchy) >= 2 else hierarchy[-1]).get("name") or "Не указано"
        module = hierarchy[-1].get("name") or "Не указано"
        return department, cfo, module

    def _catalog_name(self, collection: str, item_id: str | None) -> str:
        if not item_id:
            return ""
        item = self.repo.get_by_id(collection, item_id)
        if not item:
            return item_id
        return item["name"]

    def _category_name(self, collection: str, item_id: str | None) -> str:
        if not item_id:
            return ""
        item = self.repo.get_by_id(collection, item_id)
        if not item:
            return ""
        parent_id = item.get("parent_id")
        if not parent_id:
            return ""
        return self._catalog_name(collection, parent_id)

    def _request_items(self, request_id: str, is_income: bool | None = None) -> list[dict]:
        rows: list[dict] = []
        for item in self.repo.load_all("req_items"):
            if item.get("request_id") != request_id or item.get("status") == "deleted":
                continue
            if is_income is not None and bool(item.get("is_income", False)) != is_income:
                continue
            is_dds = bool(item.get("dds_id"))
            catalog, field, kind = ("dds_catalog", "dds_id", "ДДС") if is_dds else ("invests_catalog", "invest_id", "Инвест")
            rows.append(
                {
                    "kind": kind,
                    "purpose": "Доход" if item.get("is_income", False) else "Расход",
                    "item_id": item["id"],
                    "article": self._category_name(catalog, item.get(field)),
                    "category": self._catalog_name(catalog, item.get(field)),
                    "sum_plan": float(item.get("sum_plan") or 0),
                    "sum_fact": item.get("sum_fact"),
                    "status_code": item.get("status"),
                    "status": ITEM_STATUS_LABELS.get(item.get("status"), item.get("status") or ""),
                    "comment": item.get("comment") or "",
                    "name": item.get("name") or "",
                    "justification": item.get("justification") or "",
                    **{field: str(item.get(field) or "").strip() for field in ANALYTICS_FIELDS},
                }
            )
        return rows

    CLOSED_STATUSES = {status.value for status in EXPORTABLE_REQUEST_STATUSES} | {"rejected"}
    DEFAULT_EXPORT_STATUSES = {status.value for status in EXPORTABLE_REQUEST_STATUSES}

    def export_closed_request(self, user: dict, request_id: str) -> Path:
        request = get_required(self.repo, "requests", request_id)
        self.permissions.require_view_request(user, request)
        if request.get("status") not in self.CLOSED_STATUSES:
            raise HTTPException(status_code=400, detail="Экспорт доступен только для закрытых заявок")
        return self._write_request_workbook([request], f"request_{request_id[:8]}.xlsx")

    def export_closed_requests(
        self,
        user: dict,
        unit_id: str | None = None,
        statuses: set[str] | None = None,
        include_files: bool = False,
        *,
        department_id: str | None = None,
        department_ids: set[str] | None = None,
        module_ids: set[str] | None = None,
        fixed_only: bool = False,
        export_kind: str = "all",
        request_ids: set[str] | None = None,
    ) -> Path:
        selected_statuses = self.DEFAULT_EXPORT_STATUSES if statuses is None else statuses
        if not selected_statuses or not selected_statuses.issubset(self.CLOSED_STATUSES):
            raise HTTPException(status_code=400, detail="Выберите допустимые статусы для экспорта")
        if export_kind not in {"all", "income", "expense"}:
            raise HTTPException(status_code=400, detail="Выберите допустимый состав экспорта")
        is_income = {"income": True, "expense": False}.get(export_kind)

        selected_unit_ids = self._export_unit_ids(unit_id, department_id, department_ids, module_ids)
        position_request_ids = {
            item["request_id"]
            for item in self.repo.load_all("req_items")
            if not fixed_only or item.get("fixed")
        }
        requests = []
        for status in selected_statuses:
            for item in self.requests.list_requests(user, status=status):
                if request_ids is not None and item.get("id") not in request_ids:
                    continue
                if selected_unit_ids is not None and item.get("unit_id") not in selected_unit_ids:
                    continue
                if item.get("id") not in position_request_ids:
                    continue
                requests.append(item)
        if is_income is not None:
            requests = [item for item in requests if self._request_items(item["id"], is_income)]
        if not requests:
            raise HTTPException(status_code=404, detail="Нет закрытых заявок для экспорта")
        attachments = self._collect_export_attachments(requests, is_income) if include_files else []
        is_zgd_export = user.get("role") == "zgd"
        filename = "Заявки_ЗГД.xlsx" if is_zgd_export else {"income": "Доходы_бюджета.xlsx", "expense": "Расходы_бюджета.xlsx"}.get(export_kind, "Утверждение_бюджета.xlsx")
        workbook = (
            self._write_zgd_grouped_workbook(requests, filename, attachments, is_income)
            if is_zgd_export
            else self._write_request_workbook(requests, filename, attachments, is_income)
        )
        if not include_files:
            return workbook
        return self._write_export_archive(user, workbook, attachments)

    def _export_unit_ids(
        self,
        unit_id: str | None,
        department_id: str | None,
        department_ids: set[str] | None,
        module_ids: set[str] | None,
    ) -> set[str] | None:
        """Return the units selected by an export scope.

        ``unit_id`` remains supported for existing API consumers. Any selected
        department or nested unit includes that unit and its descendants.
        """
        selected_roots = set(department_ids or set()) | set(module_ids or set())
        if department_id:
            selected_roots.add(department_id)
        if selected_roots:
            units = {item["id"]: item for item in self.repo.load_all("units")}
            selected_units: set[str] = set()
            for selected_id in selected_roots:
                selected_units.update(self._department_unit_ids(selected_id, units))
            return selected_units
        return {unit_id} if unit_id else None

    @staticmethod
    def _department_unit_ids(department_id: str, units: dict[str, dict]) -> set[str]:
        if department_id not in units:
            raise HTTPException(status_code=404, detail="Подразделение не найдено")
        selected = {department_id}
        changed = True
        while changed:
            changed = False
            for unit in units.values():
                if unit.get("parent_id") in selected and unit["id"] not in selected:
                    selected.add(unit["id"])
                    changed = True
        return selected

    # Compat aliases
    def export_fixed_request(self, user: dict, request_id: str) -> Path:
        return self.export_closed_request(user, request_id)

    def export_fixed_requests(self, user: dict, unit_id: str | None = None) -> Path:
        return self.export_closed_requests(user, unit_id)

    def _collect_export_attachments(
        self,
        requests: list[dict],
        is_income: bool | None = None,
        item_ids: set[str] | None = None,
    ) -> list[dict]:
        request_ids = {item["id"] for item in requests}
        requests_by_id = {item["id"]: item for item in requests}
        items = {
            item["id"]: item
            for item in self.repo.load_all("req_items")
            if item.get("request_id") in request_ids
            and item.get("status") != "deleted"
            and (item_ids is None or item["id"] in item_ids)
            and (is_income is None or bool(item.get("is_income", False)) == is_income)
        }
        links = self.repo.load_all("req_item_files")
        files = {item["id"]: item for item in self.repo.load_all("files")}
        catalogs = {
            "dds": {item["id"]: item for item in self.repo.load_all("dds_catalog")},
            "invest": {item["id"]: item for item in self.repo.load_all("invests_catalog")},
        }
        attachments = []
        written: set[str] = set()
        for link in links:
                item = items.get(link.get("req_item_id"))
                file = files.get(link.get("file_id"))
                if not item or not file:
                    continue
                request = requests_by_id[item["request_id"]]
                department_name, cfo_name, module_name = self._zgd_unit_groups(request.get("unit_id"))
                department_name = self._archive_name(department_name, "Подразделение")
                cfo_name = self._archive_name(cfo_name, "Группа")
                module_name = self._archive_name(module_name, "Модуль")
                catalog = catalogs["dds"] if item.get("dds_id") else catalogs["invest"]
                article = catalog.get(item.get("dds_id") or item.get("invest_id"), {})
                article_name = self._archive_name(article.get("name"), "Статья")
                original_name = self._archive_name(file["original_name"], "Файл")
                archive_path = f"Приложения/{department_name}/{cfo_name}/{module_name}/{article_name}/{original_name}"
                duplicate_index = 2
                base_path = archive_path
                while archive_path in written:
                    archive_path = f"{base_path}_{duplicate_index}"
                    duplicate_index += 1
                written.add(archive_path)
                attachments.append(
                    {
                        "file_id": file["id"],
                        "item_id": item["id"],
                        "module_name": module_name,
                        "article_name": article_name,
                        "original_name": original_name,
                        "archive_path": archive_path,
                    }
                )
        return attachments

    def _write_export_archive(self, user: dict, workbook: Path, attachments: list[dict]) -> Path:
        archive = self.export_dir / f"{workbook.stem}.zip"

        with zipfile.ZipFile(archive, "w", compression=zipfile.ZIP_DEFLATED) as bundle:
            bundle.write(workbook, arcname=workbook.name)
            for attachment in attachments:
                body, _file, _storage, _size, _content_type = self.files.download(user, attachment["file_id"])
                try:
                    content = body.read()
                finally:
                    close = getattr(body, "close", None)
                    if callable(close):
                        close()
                bundle.writestr(attachment["archive_path"], content)
        return archive

    def _write_zgd_grouped_workbook(
        self,
        requests: list[dict],
        filename: str,
        attachments: list[dict],
        is_income: bool | None,
    ) -> Path:
        attachments_by_item: dict[str, list[dict]] = {}
        for attachment in attachments:
            attachments_by_item.setdefault(attachment["item_id"], []).append(attachment)

        item_rows: list[dict] = []
        for request in requests:
            for item in self._request_items(request["id"], is_income):
                item_rows.append(item)
        month_map = self._month_amounts_by_item(
            {item["item_id"] for item in item_rows},
            {item["item_id"]: float(item["sum_plan"] or 0) for item in item_rows},
        )

        departments: dict[str, dict] = {}
        for request in requests:
            department_name, cfo_name, module_name = self._zgd_unit_groups(request.get("unit_id"))
            request_status = REQUEST_STATUS_LABELS.get(request.get("status"), request.get("status") or "")
            for item in self._request_items(request["id"], is_income):
                article_key = f'{item["kind"]}\u0000{item["article"]}'
                planned = float(item["sum_plan"] or 0)
                approved = float(item["sum_fact"] or 0) if item["status_code"] in APPROVED_ITEM_STATUSES else 0.0
                months = month_map.get(item["item_id"], self._empty_months())

                department = departments.setdefault(
                    department_name,
                    {"planned": 0.0, "approved": 0.0, "months": self._empty_months(), "cfos": {}},
                )
                cfo = department["cfos"].setdefault(
                    cfo_name,
                    {"planned": 0.0, "approved": 0.0, "months": self._empty_months(), "articles": {}},
                )
                article = cfo["articles"].setdefault(
                    article_key,
                    {"name": item["article"], "planned": 0.0, "approved": 0.0, "months": self._empty_months(), "requests": {}},
                )
                request_row = article["requests"].setdefault(
                    request["id"],
                    {
                        "module": module_name,
                        "status": request_status,
                        "planned": 0.0,
                        "approved": 0.0,
                        "months": self._empty_months(),
                        "attachments": [],
                    },
                )
                item_attachments = attachments_by_item.get(item["item_id"], [])
                request_row["attachments"].extend(item_attachments)
                request_row["planned"] += planned
                request_row["approved"] += approved
                request_row["months"] = self._add_months(request_row["months"], months)
                article["planned"] += planned
                article["approved"] += approved
                article["months"] = self._add_months(article["months"], months)
                cfo["planned"] += planned
                cfo["approved"] += approved
                cfo["months"] = self._add_months(cfo["months"], months)
                department["planned"] += planned
                department["approved"] += approved
                department["months"] = self._add_months(department["months"], months)

        wb = Workbook()
        ws = wb.active
        ws.title = "Заявки ЗГД"
        files_column = 18
        month_start = 3
        self._style_header(
            ws,
            ["Подразделение / группа / статья / заявка", "Статус", *MONTH_HEADERS, "План", "Факт", "Корректировка", "Приложения"],
        )

        def append_row(
            name: str,
            status: str,
            planned: float,
            approved: float,
            months: list[float],
            *,
            fill: PatternFill | None = None,
            bold: bool = False,
            indent: int = 0,
            attachments_for_row: list[dict] | None = None,
        ) -> None:
            ws.append([name, status, *months, planned, approved, approved - planned, ""])
            row_number = ws.max_row
            ws.cell(row_number, 1).alignment = Alignment(indent=indent, vertical="center")
            ws.row_dimensions[row_number].outline_level = min(indent, 7)
            if fill:
                for cell in ws[row_number]:
                    cell.fill = fill
                    if bold:
                        cell.font = Font(bold=True)
            self._write_attachment_names_cell(ws.cell(row_number, files_column), attachments_for_row or [])

        for department_name in sorted(departments, key=str.casefold):
            department = departments[department_name]
            append_row(department_name, "", department["planned"], department["approved"], department["months"], fill=DEPARTMENT_FILL, bold=True)
            for cfo_name in sorted(department["cfos"], key=str.casefold):
                cfo = department["cfos"][cfo_name]
                append_row(cfo_name, "", cfo["planned"], cfo["approved"], cfo["months"], fill=CFO_FILL, bold=True, indent=1)
                articles = cfo["articles"]
                for article_key in sorted(articles, key=lambda value: articles[value]["name"].casefold()):
                    article = articles[article_key]
                    append_row(article["name"], "", article["planned"], article["approved"], article["months"], fill=ARTICLE_FILL, bold=True, indent=2)
                    for request_id, request_row in sorted(article["requests"].items(), key=lambda value: value[1]["module"].casefold()):
                        append_row(
                            request_row["module"],
                            request_row["status"],
                            request_row["planned"],
                            request_row["approved"],
                            request_row["months"],
                            indent=3,
                            attachments_for_row=request_row["attachments"],
                        )

        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 30
        self._format_money_columns(ws, range(month_start, month_start + 15))
        self._autosize(ws)
        self._hide_columns(ws, month_start, 12)
        self._enable_sheet_grouping(ws)
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(files_column)}{ws.max_row}"
        ws.freeze_panes = "A2"

        details = wb.create_sheet("Детализация заявок")
        max_attachments = max((len(items) for items in attachments_by_item.values()), default=0)
        attachment_headers = [f"Приложение {index}" for index in range(1, max_attachments + 1)]
        details_columns = [
                "Подразделение",
                "Группа",
                "Модуль",
                "Статус заявки",
                "Тип",
                "Назначение",
                "Категория",
                "Статья / проект",
                "Наименование",
                "Обоснование",
                *MONTH_HEADERS,
                "План",
                "Факт",
                "Корректировка",
                "Статус строки",
                "Комментарий",
                *ANALYTICS_HEADERS,
                *attachment_headers,
            ]
        self._style_header(details, details_columns)
        month_start = 11
        attachment_start = len(details_columns) - max_attachments + 1
        exported_items: list[dict] = []
        for request in requests:
            department_name, cfo_name, module_name = self._zgd_unit_groups(request.get("unit_id"))
            request_status = REQUEST_STATUS_LABELS.get(request.get("status"), request.get("status") or "")
            items = self._request_items(request["id"], is_income)
            if not items:
                details.append(
                    [
                        department_name,
                        cfo_name,
                        module_name,
                        request_status,
                        "",
                        "",
                        "",
                        "Строки отсутствуют",
                        "",
                        "",
                        *self._empty_months(),
                        0,
                        0,
                        0,
                        "",
                        "",
                        *[""] * len(ANALYTICS_FIELDS),
                        *([""] * max_attachments),
                    ],
                )
                continue
            for item in items:
                approved = float(item["sum_fact"] or 0) if item["status_code"] in APPROVED_ITEM_STATUSES else 0.0
                exported_items.append({**item, "cfo_name": cfo_name, "sum_fact": approved})
                row_attachments = attachments_by_item.get(item["item_id"], [])
                details.append(
                    [
                        department_name,
                        cfo_name,
                        module_name,
                        request_status,
                        item["kind"],
                        item["purpose"],
                        item["category"],
                        item["article"],
                        item["name"],
                        item["justification"],
                        *month_map.get(item["item_id"], self._empty_months()),
                        item["sum_plan"],
                        approved,
                        approved - float(item["sum_plan"] or 0),
                        item["status"],
                        item["comment"],
                        *[item.get(field) or "" for field in ANALYTICS_FIELDS],
                        *[attachment["original_name"] for attachment in row_attachments],
                        *([""] * (max_attachments - len(row_attachments))),
                    ],
                )
                for index, attachment in enumerate(row_attachments, start=attachment_start):
                    cell = details.cell(details.max_row, index)
                    cell.hyperlink = attachment["archive_path"]
                    cell.style = "Hyperlink"
        self._format_money_columns(details, range(month_start, month_start + 15))
        self._autosize(details)
        self._hide_columns(details, month_start, 12)
        if details.max_row > 1:
            details.auto_filter.ref = f"A1:{get_column_letter(len(details_columns))}{details.max_row}"
        details.freeze_panes = "A2"
        self._prepend_request_summary_sheets(
            wb,
            requests,
            exported_items,
            plan_key="sum_plan",
            fact_key="sum_fact",
            cfo_name="cfo_name",
        )

        target = self.export_dir / filename
        wb.save(target)
        return target

    @staticmethod
    def _archive_name(value: Any, fallback: str) -> str:
        name = re.sub(r'[\\/:*?"<>|\x00-\x1f]+', "_", str(value or "").strip()).strip(". ")
        return name or fallback

    def _write_request_workbook(
        self,
        requests: list[dict],
        filename: str,
        attachments: list[dict] | None = None,
        is_income: bool | None = None,
    ) -> Path:
        wb = Workbook()
        attachments_by_item: dict[str, list[dict]] = {}
        for attachment in attachments or []:
            attachments_by_item.setdefault(attachment["item_id"], []).append(attachment)
        max_attachments = max((len(items) for items in attachments_by_item.values()), default=0)
        attachment_headers = [f"Приложение {index}" for index in range(1, max_attachments + 1)]
        exported_items: list[dict] = []

        composition = wb.active
        composition.title = "Состав"
        composition_columns = [
                "Подразделение",
                "Модуль",
                "Статус заявки",
                "Тип",
                "Назначение",
                "Категория",
                "Статья / проект",
                "Наименование",
                "Обоснование",
                *MONTH_HEADERS,
                "План",
                "Факт",
                "Статус строки",
                "Комментарий",
                *ANALYTICS_HEADERS,
                *attachment_headers,
            ]
        self._style_header(composition, composition_columns)
        month_start = 10
        attachment_start = len(composition_columns) - max_attachments + 1
        all_items = [
            item
            for request in requests
            for item in self._request_items(request["id"], is_income)
        ]
        month_map = self._month_amounts_by_item(
            {item["item_id"] for item in all_items},
            {item["item_id"]: float(item["sum_plan"] or 0) for item in all_items},
        )
        for request in requests:
            module_name = self._unit_name(request.get("unit_id"))
            department_name = self._department_name(request.get("unit_id"))
            request_status = REQUEST_STATUS_LABELS.get(request.get("status"), request.get("status") or "")
            items = self._request_items(request["id"], is_income)
            _department, cfo_name, _module = self._zgd_unit_groups(request.get("unit_id"))
            exported_items.extend(
                {
                    **item,
                    "cfo_name": cfo_name,
                    "sum_fact": float(item["sum_fact"] or 0),
                }
                for item in items
            )
            if not items:
                composition.append(
                    [
                        department_name,
                        module_name,
                        request_status,
                        "",
                        "",
                        "",
                        "Строки отсутствуют",
                        "",
                        "",
                        *self._empty_months(),
                        0,
                        None,
                        "",
                        "",
                        *[""] * len(ANALYTICS_FIELDS),
                        *([""] * max_attachments),
                    ]
                )
                continue
            for item in items:
                row_attachments = attachments_by_item.get(item["item_id"], [])
                composition.append(
                    [
                        department_name,
                        module_name,
                        request_status,
                        item["kind"],
                        item["purpose"],
                        item["category"],
                        item["article"],
                        item["name"],
                        item["justification"],
                        *month_map.get(item["item_id"], self._empty_months()),
                        item["sum_plan"],
                        item["sum_fact"],
                        item["status"],
                        item["comment"],
                        *[item.get(field) or "" for field in ANALYTICS_FIELDS],
                        *[attachment["original_name"] for attachment in row_attachments],
                        *([""] * (max_attachments - len(row_attachments))),
                    ]
                )
                for index, attachment in enumerate(row_attachments, start=attachment_start):
                    file_cell = composition.cell(composition.max_row, index)
                    file_cell.hyperlink = attachment["archive_path"]
                    file_cell.style = "Hyperlink"
        self._format_money_columns(composition, range(month_start, month_start + 14))
        self._autosize(composition)
        self._hide_columns(composition, month_start, 12)
        if composition.max_row > 1:
            composition.auto_filter.ref = f"A1:{get_column_letter(len(composition_columns))}{composition.max_row}"
        composition.freeze_panes = "A2"
        self._prepend_request_summary_sheets(
            wb,
            requests,
            exported_items,
            plan_key="sum_plan",
            fact_key="sum_fact",
            cfo_name="cfo_name",
        )

        target = self.export_dir / filename
        wb.save(target)
        return target

    def export_approval_register(
        self,
        user: dict,
        *,
        view: str = "cfo",
        group_by: list[str] | None = None,
        item_ids: set[str] | None = None,
        include_files: bool = True,
        export_kind: str = "all",
        fixed_only: bool = False,
        department_ids: set[str] | None = None,
        module_ids: set[str] | None = None,
        **filters,
    ) -> Path:
        if export_kind not in {"all", "income", "expense"}:
            raise HTTPException(status_code=400, detail="Выберите допустимый состав экспорта")
        is_income = {"income": True, "expense": False}.get(export_kind)
        requested_flow = filters.pop("is_income", None)
        # The explicit export setting must win over the current table filter.
        # This lets a user switch from the visible type to another one directly
        # in the dialog instead of receiving a confusing 400 response.
        if requested_flow is not None and is_income is None:
            is_income = requested_flow
        selected_unit_ids = self._export_unit_ids(None, None, department_ids, module_ids)
        payload = self.requests.approval_register(
            user,
            view,
            group_by=group_by,
            item_ids=item_ids,
            is_income=is_income,
            fixed_only=fixed_only,
            module_ids=selected_unit_ids,
            **filters,
        )
        notes_filters = dict(filters)
        if item_ids is not None:
            notes_filters["item_ids"] = item_ids
        if export_kind != "all":
            notes_filters["export_kind"] = export_kind
        if fixed_only:
            notes_filters["fixed_only"] = True
        items = payload.get("summary_items") or []
        requests = []
        seen: set[str] = set()
        for item in items:
            request_id = item.get("request_id")
            if not request_id or request_id in seen:
                continue
            seen.add(request_id)
            requests.append({"id": request_id, "unit_id": item.get("module_id")})
        attachments = (
            self._collect_export_attachments(requests, is_income, item_ids={item["id"] for item in items})
            if include_files and requests else []
        )
        workbook = self._write_register_workbook(
            payload,
            "Табличный_вид.xlsx",
            filters=notes_filters,
            attachments=attachments,
        )
        if not attachments:
            return workbook
        return self._write_export_archive(user, workbook, attachments)

    def _write_register_workbook(
        self,
        payload: dict,
        filename: str,
        *,
        filters: dict | None = None,
        attachments: list[dict] | None = None,
    ) -> Path:
        aggregates = payload.get("aggregates") or {}
        groups = payload.get("groups") or []
        items = payload.get("summary_items") or []
        analytics_summary = payload.get("analytics_summary") or []
        grouping = payload.get("group_by") or []
        view = payload.get("view") or "cfo"
        attachments_by_item: dict[str, list[dict]] = {}
        for attachment in attachments or []:
            attachments_by_item.setdefault(attachment["item_id"], []).append(attachment)
        max_attachments = max((len(files) for files in attachments_by_item.values()), default=0)
        attachment_headers = [f"Приложение {index}" for index in range(1, max_attachments + 1)]
        month_map = self._month_amounts_by_item(
            {item["id"] for item in items},
            {item["id"]: float(item.get("requested_sum") or 0) for item in items},
        )

        def months_for_scope(scope: dict) -> list[float]:
            totals = self._empty_months()
            if not scope:
                return totals
            for item in items:
                if not self._item_matches_group_scope(item, scope):
                    continue
                totals = self._add_months(totals, month_map.get(item["id"], self._empty_months()))
            return totals

        wb = Workbook()
        summary = wb.active
        summary.title = "Сводка"
        total_rows = int(aggregates.get("total_rows") or 0)
        approved_rows = int(aggregates.get("approved_rows") or 0)
        rejected_rows = int(aggregates.get("rejected_rows") or 0)
        readiness = 0 if not total_rows else round((approved_rows + rejected_rows) / total_rows * 100)
        grouping_label = " → ".join(REGISTER_GROUP_LABELS.get(level, level) for level in grouping) or "—"
        notes = [
            f"Представление: {REGISTER_VIEW_LABELS.get(view, view)}",
            f"Группировка: {grouping_label}",
            *self._register_filter_notes(filters or {}),
        ]
        if attachments:
            notes.append(f"Файлов в архиве: {len(attachments)}")
        self._fill_kpi_sheet(
            summary,
            [
                ("Всего строк", total_rows, "int"),
                ("Запрошено", float(aggregates.get("requested_sum") or 0), "money"),
                ("Согласовано", float(aggregates.get("approved_sum") or 0), "money"),
                ("Корректировка", float(aggregates.get("difference") or 0), "money"),
                ("На рассмотрении", float(aggregates.get("pending_sum") or 0), "money"),
                ("% согласования", readiness, "int"),
                ("Заявок", int(aggregates.get("requests_count") or 0), "int"),
                ("Файлов", len(attachments or []), "int"),
            ],
            notes=notes,
        )

        analytics_sheet = wb.create_sheet("Сводка по аналитикам")
        self._fill_register_analytics_sheet(analytics_sheet, analytics_summary)

        registry = wb.create_sheet("Реестр")
        registry_columns = [
            "Структура",
            "Уровень",
            *MONTH_HEADERS,
            "План, ₽",
            "Факт, ₽",
            "Корректировка, ₽",
            "Статус",
            "Строк",
            *ANALYTICS_HEADERS,
            "Файлы",
        ]
        self._style_header(registry, registry_columns)
        used_item_ids: set[str] = set()
        files_column = len(registry_columns)
        month_start = 3

        def write_registry_row(
            values: list[Any],
            *,
            level: int,
            fill: PatternFill | None = None,
            bold: bool = False,
            row_attachments: list[dict] | None = None,
        ) -> None:
            registry.append(values)
            row_number = registry.max_row
            registry.row_dimensions[row_number].outline_level = min(level, 7)
            registry.cell(row_number, 1).alignment = Alignment(indent=level, vertical="center")
            if fill:
                for cell in registry[row_number]:
                    cell.fill = fill
                    if bold:
                        cell.font = Font(bold=True)
            self._write_attachment_names_cell(registry.cell(row_number, files_column), row_attachments or [])

        def append_group(group: dict, level: int = 0) -> None:
            analytics_fields = (group.get("analytics") or {}).get("fields") or {}
            aggregates_row = group.get("aggregates") or {}
            children = group.get("children") or []
            write_registry_row(
                [
                    group.get("name") or "",
                    group.get("label") or REGISTER_GROUP_LABELS.get(group.get("type"), group.get("type") or ""),
                    *months_for_scope(group.get("scope") or {}),
                    float(aggregates_row.get("requested_sum") or 0),
                    float(aggregates_row.get("approved_sum") or 0),
                    float(aggregates_row.get("difference") or 0),
                    AGGREGATE_STATUS_LABELS.get(aggregates_row.get("aggregate_status"), aggregates_row.get("aggregate_status") or ""),
                    int(aggregates_row.get("total_rows") or 0),
                    *[self._group_analytics_value(analytics_fields.get(field)) for field in ANALYTICS_FIELDS],
                    "",
                ],
                level=level,
                fill=GROUP_FILLS[min(level, len(GROUP_FILLS) - 1)],
                bold=True,
            )
            for child in children:
                append_group(child, level + 1)
            if children:
                return
            scope = group.get("scope") or {}
            if not scope:
                return
            for item in items:
                if item["id"] in used_item_ids or not self._item_matches_group_scope(item, scope):
                    continue
                used_item_ids.add(item["id"])
                planned = float(item.get("requested_sum") or 0)
                approved = float(item.get("approved_sum") or 0)
                row_attachments = attachments_by_item.get(item["id"], [])
                write_registry_row(
                    [
                        item.get("name") or "",
                        "Строка",
                        *month_map.get(item["id"], self._empty_months()),
                        planned,
                        approved,
                        approved - planned,
                        ITEM_STATUS_LABELS.get(item.get("status"), item.get("status") or ""),
                        1,
                        *[item.get(field) or "" for field in ANALYTICS_FIELDS],
                        "",
                    ],
                    level=level + 1,
                    row_attachments=row_attachments,
                )

        for group in groups:
            append_group(group)
        self._format_money_columns(registry, range(month_start, month_start + 15))
        self._autosize(registry)
        self._hide_columns(registry, month_start, 12)
        self._enable_sheet_grouping(registry)
        registry.freeze_panes = "A2"

        details = wb.create_sheet("Строки")
        details_columns = [
            "ЦФО",
            "Модуль",
            "Статья",
            "Категория",
            "Наименование",
            "Статус заявки",
            "Статус строки",
            *MONTH_HEADERS,
            "План, ₽",
            "Факт, ₽",
            "Корректировка, ₽",
            "Обоснование",
            "Комментарий",
            *ANALYTICS_HEADERS,
            "Год бюджета",
            "Тип",
            *attachment_headers,
        ]
        self._style_header(details, details_columns)
        details_month_start = 8
        attachment_start = len(details_columns) - max_attachments + 1
        for item in items:
            planned = float(item.get("requested_sum") or 0)
            approved = float(item.get("approved_sum") or 0)
            row_attachments = attachments_by_item.get(item["id"], [])
            details.append(
                [
                    item.get("cfo_name") or "",
                    item.get("module_name") or "",
                    item.get("article_name") or "",
                    item.get("category_name") or "",
                    item.get("name") or "",
                    REQUEST_STATUS_LABELS.get(item.get("request_status"), item.get("request_status") or ""),
                    ITEM_STATUS_LABELS.get(item.get("status"), item.get("status") or ""),
                    *month_map.get(item["id"], self._empty_months()),
                    planned,
                    approved,
                    approved - planned,
                    item.get("justification") or "",
                    item.get("comment") or "",
                    *[item.get(field) or "" for field in ANALYTICS_FIELDS],
                    item.get("budget_year") or "",
                    "Инвест" if item.get("kind") == "invest" else "ДДС",
                    *[attachment["original_name"] for attachment in row_attachments],
                    *([""] * (max_attachments - len(row_attachments))),
                ]
            )
            for index, attachment in enumerate(row_attachments, start=attachment_start):
                cell = details.cell(details.max_row, index)
                cell.hyperlink = attachment["archive_path"]
                cell.style = "Hyperlink"
        self._format_money_columns(details, range(details_month_start, details_month_start + 15))
        self._autosize(details)
        self._hide_columns(details, details_month_start, 12)
        if details.max_row > 1:
            details.auto_filter.ref = f"A1:{details.cell(1, len(details_columns)).column_letter}{details.max_row}"
        details.freeze_panes = "A2"

        target = self.export_dir / filename
        wb.save(target)
        return target

    @staticmethod
    def _item_matches_group_scope(item: dict, scope: dict) -> bool:
        for key, expected in scope.items():
            actual = item.get(key)
            if key in ANALYTICS_FIELDS:
                actual = str(actual or "").strip()
                expected_value = "" if expected in {None, "", "__empty__"} else str(expected)
                if actual != expected_value:
                    return False
                continue
            if str(actual or "") != str(expected or ""):
                return False
        return True

    @staticmethod
    def _write_attachment_names_cell(cell, attachments: list[dict]) -> None:
        if not attachments:
            cell.value = ""
            return
        seen: set[str] = set()
        names = []
        for attachment in attachments:
            name = attachment.get("original_name") or ""
            if not name or name in seen:
                continue
            seen.add(name)
            names.append(name)
        cell.value = "\n".join(names)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        archive_path = attachments[0].get("archive_path")
        if archive_path:
            cell.hyperlink = archive_path
            cell.style = "Hyperlink"

    @staticmethod
    def _enable_sheet_grouping(ws) -> None:
        outline = ws.sheet_properties.outlinePr
        if outline is None:
            outline = Outline()
            ws.sheet_properties.outlinePr = outline
        outline.summaryBelow = False
        outline.summaryRight = False
        max_level = 0
        for row in range(2, (ws.max_row or 1) + 1):
            level = int(ws.row_dimensions[row].outline_level or 0)
            if level > max_level:
                max_level = level
        ws.sheet_format.outlineLevelRow = max_level

    @staticmethod
    def _group_analytics_value(field: dict | None) -> str:
        if not field:
            return ""
        if field.get("mixed"):
            return "Разные значения"
        return str(field.get("value") or "")

    def _register_filter_notes(self, filters: dict) -> list[str]:
        notes: list[str] = []
        search = str(filters.get("search") or "").strip()
        if search:
            notes.append(f"Поиск: {search}")
        status = filters.get("status")
        if status:
            notes.append(f"Статус строки: {ITEM_STATUS_LABELS.get(status, status)}")
        request_status = filters.get("request_status")
        if request_status:
            labels = [
                REQUEST_STATUS_LABELS.get(part.strip(), part.strip())
                for part in str(request_status).split(",")
                if part.strip()
            ]
            if labels:
                notes.append(f"Статус заявки: {', '.join(labels)}")
        budget_year = filters.get("budget_year")
        if budget_year:
            notes.append(f"Год: {budget_year}")
        if filters.get("cfo_id"):
            notes.append(f"ЦФО: {self._unit_name(filters.get('cfo_id')) or filters.get('cfo_id')}")
        if filters.get("article_id"):
            notes.append(
                f"Статья: {self._catalog_name('dds_catalog', filters.get('article_id')) or self._catalog_name('invests_catalog', filters.get('article_id')) or filters.get('article_id')}"
            )
        if filters.get("module_id"):
            notes.append(f"Модуль: {self._unit_name(filters.get('module_id')) or filters.get('module_id')}")
        if filters.get("request_id"):
            notes.append(f"Заявка: {filters.get('request_id')}")
        export_kind = filters.get("export_kind")
        if export_kind == "income":
            notes.append("Состав: только доходы")
        elif export_kind == "expense":
            notes.append("Состав: только расходы")
        if filters.get("fixed_only"):
            notes.append("Только зафиксированные строки")
        for field in ANALYTICS_FIELDS:
            value = str(filters.get(field) or "").strip()
            if value:
                label = REGISTER_GROUP_LABELS.get(field, field)
                notes.append(f"{label}: {'Не заполнено' if value == '__empty__' else value}")
        if filters.get("item_ids") is not None:
            notes.append("Применены фильтры колонок таблицы")
        return notes

    def _fill_kpi_sheet(
        self,
        ws,
        metrics: list[tuple[str, Any, str]],
        *,
        notes: list[str] | None = None,
    ) -> None:
        self._style_header(ws, ["Показатель", "Значение"])
        for label, value, kind in metrics:
            ws.append([label, value])
            cell = ws.cell(ws.max_row, 2)
            if kind == "money":
                cell.number_format = MONEY_FORMAT
            elif kind == "percent":
                cell.number_format = '0"%"'
            ws.cell(ws.max_row, 1).fill = SUMMARY_FILL
            cell.fill = SUMMARY_FILL
        if notes:
            ws.append([])
            ws.append(["Параметры выгрузки", ""])
            ws.cell(ws.max_row, 1).font = Font(bold=True)
            for note in notes:
                ws.append([note, ""])
                ws.cell(ws.max_row, 1).font = NOTE_FONT
        self._autosize(ws)
        ws.freeze_panes = "A2"

    def _fill_register_analytics_sheet(self, ws, analytics_summary: list[dict]) -> None:
        self._style_header(
            ws,
            [
                "Поле",
                "Значение",
                "План, ₽",
                "Согласовано, ₽",
                "Строк",
                "Наибольшая нагрузка ЦФО",
                "План ЦФО, ₽",
                "Строк ЦФО",
            ],
        )
        if not analytics_summary:
            ws.append(["Нет заполненных аналитик", "", "", "", "", "", "", ""])
            self._autosize(ws)
            return
        for section in analytics_summary:
            label = section.get("label") or section.get("field") or ""
            values = section.get("values") or []
            ws.append([label, f"{len(values)} значений", "", "", "", "", "", ""])
            header_row = ws.max_row
            ws.row_dimensions[header_row].outline_level = 0
            for cell in ws[header_row]:
                cell.fill = SUMMARY_FILL
                cell.font = Font(bold=True)
            for value in values:
                aggregates = value.get("aggregates") or {}
                top_cfo = value.get("top_cfo") or {}
                ws.append(
                    [
                        label,
                        value.get("value") or "",
                        float(aggregates.get("requested_sum") or 0),
                        float(aggregates.get("approved_sum") or 0),
                        int(aggregates.get("total_rows") or 0),
                        top_cfo.get("cfo_name") or "",
                        float(top_cfo.get("requested_sum") or 0),
                        int(top_cfo.get("total_rows") or 0),
                    ]
                )
                ws.row_dimensions[ws.max_row].outline_level = 1
                for column in (3, 4, 7):
                    ws.cell(ws.max_row, column).number_format = MONEY_FORMAT
        self._autosize(ws)
        self._enable_sheet_grouping(ws)
        ws.freeze_panes = "A2"

    def _fill_item_analytics_sheet(self, ws, items: list[dict], *, plan_key: str, fact_key: str, cfo_name: str) -> None:
        grouped: dict[str, dict[str, dict]] = {field: {} for field in ANALYTICS_FIELDS}
        for item in items:
            cfo = str(item.get(cfo_name) or "ЦФО не указан")
            planned = float(item.get(plan_key) or 0)
            approved = float(item.get(fact_key) or 0)
            for field in ANALYTICS_FIELDS:
                value = str(item.get(field) or "").strip()
                if not value:
                    continue
                bucket = grouped[field].setdefault(
                    value,
                    {
                        "planned": 0.0,
                        "approved": 0.0,
                        "rows": 0,
                        "cfos": {},
                    },
                )
                bucket["planned"] += planned
                bucket["approved"] += approved
                bucket["rows"] += 1
                cfo_load = bucket["cfos"].setdefault(cfo, {"planned": 0.0, "rows": 0})
                cfo_load["planned"] += planned
                cfo_load["rows"] += 1
        sections = []
        for field in ANALYTICS_FIELDS:
            values = grouped[field]
            if not values:
                continue
            rows = []
            for value, bucket in values.items():
                top_cfo_name, top_cfo = max(
                    bucket["cfos"].items(),
                    key=lambda entry: (entry[1]["planned"], entry[1]["rows"], entry[0].casefold()),
                )
                rows.append(
                    {
                        "value": value,
                        "aggregates": {
                            "requested_sum": bucket["planned"],
                            "approved_sum": bucket["approved"],
                            "total_rows": bucket["rows"],
                        },
                        "top_cfo": {
                            "cfo_name": top_cfo_name,
                            "requested_sum": top_cfo["planned"],
                            "total_rows": top_cfo["rows"],
                        },
                    }
                )
            sections.append(
                {
                    "field": field,
                    "label": REGISTER_GROUP_LABELS[field],
                    "values": sorted(rows, key=lambda item: (-item["aggregates"]["requested_sum"], item["value"].casefold())),
                }
            )
        self._fill_register_analytics_sheet(ws, sections)

    def _prepend_request_summary_sheets(
        self,
        wb: Workbook,
        requests: list[dict],
        items: list[dict],
        *,
        plan_key: str,
        fact_key: str,
        cfo_name: str,
    ) -> None:
        planned = sum(float(item.get(plan_key) or 0) for item in items)
        approved = sum(float(item.get(fact_key) or 0) for item in items)
        summary = wb.create_sheet("Сводка", 0)
        self._fill_kpi_sheet(
            summary,
            [
                ("Заявок", len(requests), "int"),
                ("Строк", len(items), "int"),
                ("План", planned, "money"),
                ("Факт", approved, "money"),
                ("Корректировка", approved - planned, "money"),
            ],
        )
        analytics_sheet = wb.create_sheet("Сводка по аналитикам", 1)
        self._fill_item_analytics_sheet(analytics_sheet, items, plan_key=plan_key, fact_key=fact_key, cfo_name=cfo_name)
