import io

from openpyxl import load_workbook

from app.seed import DDS_LICENSE_ID, MODULE_ALPHA_ID
from tests.test_api import auth, make_client


def test_register_and_export_can_be_limited_to_income_lines(tmp_path):
    client = make_client(tmp_path)
    employee = auth(client, "employee", "employee")
    request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()

    income = client.post(
        f"/requests/{request['id']}/items",
        json={"dds_id": DDS_LICENSE_ID, "is_income": True, "name": "Доход", "sum_plan": 120},
        headers=employee,
    )
    expense = client.post(
        f"/requests/{request['id']}/items",
        json={"dds_id": DDS_LICENSE_ID, "is_income": False, "name": "Расход", "sum_plan": 80},
        headers=employee,
    )
    assert income.status_code == 200
    assert expense.status_code == 200

    register = client.get(
        "/approval-register",
        params={"request_id": request["id"], "is_income": "true"},
        headers=employee,
    )
    assert register.status_code == 200, register.text
    assert [item["name"] for item in register.json()["summary_items"]] == ["Доход"]

    exported = client.get(
        "/approval-register/export",
        params={"request_id": request["id"], "is_income": "true"},
        headers=employee,
    )
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(io.BytesIO(exported.content))
    names = [row[4].value for row in workbook["Строки"].iter_rows(min_row=2)]
    assert names == ["Доход"]

    # A choice made in the export dialog wins over the currently displayed
    # type, so switching to a different export type does not produce a 400.
    switched = client.get(
        "/approval-register/export",
        params={"request_id": request["id"], "is_income": "true", "export_kind": "expense"},
        headers=employee,
    )
    assert switched.status_code == 200, switched.text
    workbook = load_workbook(io.BytesIO(switched.content))
    names = [row[4].value for row in workbook["Строки"].iter_rows(min_row=2)]
    assert names == ["Расход"]
