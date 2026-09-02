import io
import zipfile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.seed import CFO_ID, DDS_LICENSE_ID, MODULE_ALPHA_ID, REQUEST_ID
from tests.test_api import auth, make_client


def create_fixed_position_data(client):
    repo = client.app.state.repo
    request = repo.create(
        "requests",
        {
            "id": REQUEST_ID,
            "created_by_id": "00000000-0000-0000-0000-000000000003",
            "budget_year": 2026,
            "unit_id": MODULE_ALPHA_ID,
            "sum_plan": 100,
            "sum_fact": 80,
            "status": "approved",
        },
    )
    position = repo.create(
        "cfo_positions",
        {
            "budget_year": 2026,
            "cfo_unit_id": CFO_ID,
            "dds_id": DDS_LICENSE_ID,
            "invest_id": None,
            "is_income": False,
            "status": "approved",
            "current_step_id": None,
        },
    )
    item = repo.create(
        "req_items",
        {
            "request_id": request["id"],
            "cfo_position_id": position["id"],
            "dds_id": DDS_LICENSE_ID,
            "invest_id": None,
            "name": "Лицензия",
            "sum_plan": 100,
            "sum_fact": 80,
            "justification": "",
            "status": "approved_with_changes",
            "comment": "Снижено",
            "is_income": False,
            "frozen": True,
            "fixed": True,
            "analytics_1": "Проект А",
            "analytics_2": "",
        },
    )
    return request, position, item


def test_export_fixed_scope_uses_fixed_request_lines(tmp_path):
    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")
    create_fixed_position_data(client)

    exported = client.get(
        "/requests/export/closed",
        params={"statuses": "approved", "fixed_only": "true"},
        headers=admin,
    )
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(io.BytesIO(exported.content))
    assert "Состав" in workbook.sheetnames
    assert any(
        cell.value == "Лицензия"
        for row in workbook["Состав"].iter_rows()
        for cell in row
    )


def test_export_archive_keeps_files_linked_to_position_contributions(tmp_path):
    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")
    _request, _position, item = create_fixed_position_data(client)
    repo = client.app.state.repo
    storage = repo.create(
        "storage_objects",
        {
            "storage_bucket": "budgetbasket-files",
            "storage_key": "tests/evidence.png",
            "content_sha256": "0" * 64,
            "mime_type": "image/png",
            "size_bytes": 4,
        },
    )
    file = repo.create(
        "files",
        {"id_storage_object": storage["id"], "original_name": "evidence.png"},
    )
    repo.insert("req_item_files", {"file_id": file["id"], "req_item_id": item["id"]})
    client.app.state.file_service.object_storage.put_object(
        "tests/evidence.png", b"test", "image/png"
    )

    exported = client.get(
        "/requests/export/closed",
        params={
            "statuses": "approved",
            "fixed_only": "true",
            "include_files": "true",
        },
        headers=admin,
    )
    assert exported.status_code == 200
    with zipfile.ZipFile(io.BytesIO(exported.content)) as archive:
        names = archive.namelist()
    assert any(name.endswith(".xlsx") for name in names)
    assert any(name.endswith("evidence.png") for name in names)


def test_closed_export_includes_summary_and_analytics(tmp_path):
    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")
    create_fixed_position_data(client)

    exported = client.get(
        "/requests/export/closed",
        params={"statuses": "approved", "fixed_only": "true"},
        headers=admin,
    )
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(io.BytesIO(exported.content))
    assert workbook.sheetnames[0] == "Сводка"
    assert "Сводка по аналитикам" in workbook.sheetnames
    assert "Состав" in workbook.sheetnames
    summary_values = {row[0].value: row[1].value for row in workbook["Сводка"].iter_rows(min_row=2, max_col=2) if row[0].value}
    assert summary_values["Заявок"] == 1
    assert summary_values["Строк"] == 1
    assert summary_values["План"] == 100
    headers = [cell.value for cell in workbook["Состав"][1]]
    assert "Аналитика 1" in headers
    assert "ID заявки" not in headers
    assert headers.index("Январь") < headers.index("План")
    january = headers.index("Январь") + 1
    assert workbook["Состав"].column_dimensions[get_column_letter(january)].hidden
    analytics_index = headers.index("Аналитика 1")
    assert any(row[analytics_index].value == "Проект А" for row in workbook["Состав"].iter_rows(min_row=2))
    analytics_values = [row[1].value for row in workbook["Сводка по аналитикам"].iter_rows(min_row=2)]
    assert "Проект А" in analytics_values


def test_closed_export_request_ids_keep_visible_table_scope(tmp_path):
    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")
    visible, _position, _item = create_fixed_position_data(client)
    hidden = client.app.state.repo.create(
        "requests",
        {
            "created_by_id": "00000000-0000-0000-0000-000000000003",
            "budget_year": 2026,
            "unit_id": MODULE_ALPHA_ID,
            "sum_plan": 50,
            "sum_fact": 50,
            "status": "approved",
        },
    )
    client.app.state.repo.create(
        "req_items",
        {
            "request_id": hidden["id"],
            "cfo_position_id": None,
            "dds_id": DDS_LICENSE_ID,
            "invest_id": None,
            "name": "Скрытая строка",
            "sum_plan": 50,
            "sum_fact": 50,
            "justification": "",
            "status": "approved",
            "comment": "",
            "is_income": False,
            "frozen": True,
            "fixed": True,
        },
    )

    exported = client.get(
        "/requests/export/closed",
        params={"statuses": "approved", "request_ids": visible["id"]},
        headers=admin,
    )
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(io.BytesIO(exported.content))
    names = {cell.value for row in workbook["Состав"].iter_rows() for cell in row}
    assert "Лицензия" in names
    assert "Скрытая строка" not in names


def test_register_export_matches_filters_and_keeps_summary_analytics(tmp_path):
    client = make_client(tmp_path)
    employee = auth(client, "employee", "employee")
    request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    tagged = client.post(
        f"/requests/{request['id']}/items",
        json={"dds_id": DDS_LICENSE_ID, "name": "Tagged line", "sum_plan": 100, "analytics_1": "Проект А"},
        headers=employee,
    )
    other = client.post(
        f"/requests/{request['id']}/items",
        json={"dds_id": DDS_LICENSE_ID, "name": "Other line", "sum_plan": 40, "analytics_1": "Проект Б"},
        headers=employee,
    )
    assert tagged.status_code == 200
    assert other.status_code == 200

    exported = client.get(
        "/approval-register/export",
        params={"analytics_1": "Проект А", "view": "cfo", "request_id": request["id"]},
        headers=employee,
    )
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(io.BytesIO(exported.content))
    assert workbook.sheetnames[:4] == ["Сводка", "Сводка по аналитикам", "Реестр", "Строки"]
    summary_values = {row[0].value: row[1].value for row in workbook["Сводка"].iter_rows(min_row=2, max_col=2) if row[0].value}
    assert summary_values["Всего строк"] == 1
    assert summary_values["Запрошено"] == 100
    analytics_values = [row[1].value for row in workbook["Сводка по аналитикам"].iter_rows(min_row=2)]
    assert "Проект А" in analytics_values
    assert "Проект Б" not in analytics_values
    line_names = [row[4].value for row in workbook["Строки"].iter_rows(min_row=2)]
    assert "Tagged line" in line_names
    assert "Other line" not in line_names
    details_headers = [cell.value for cell in workbook["Строки"][1]]
    assert "ID заявки" not in details_headers
    assert details_headers.index("Январь") < details_headers.index("План, ₽")
    january = details_headers.index("Январь") + 1
    assert workbook["Строки"].column_dimensions[get_column_letter(january)].hidden
    tagged_row = next(row for row in workbook["Строки"].iter_rows(min_row=2) if row[4].value == "Tagged line")
    assert tagged_row[january - 1].value == 8.34
    registry_headers = [cell.value for cell in workbook["Реестр"][1]]
    assert registry_headers.index("Январь") < registry_headers.index("План, ₽")
    assert workbook["Реестр"].column_dimensions[get_column_letter(registry_headers.index("Январь") + 1)].hidden
    structure_names = [row[0].value for row in workbook["Реестр"].iter_rows(min_row=2)]
    assert "Tagged line" in structure_names
    assert any(value and "Tagged" not in str(value) for value in structure_names)
    registry = workbook["Реестр"]
    assert registry.sheet_properties.outlinePr.summaryBelow is False
    outline_levels = [int(registry.row_dimensions[index].outline_level or 0) for index in range(2, registry.max_row + 1)]
    assert max(outline_levels) >= 1
    analytics_sheet = workbook["Сводка по аналитикам"]
    assert analytics_sheet.sheet_properties.outlinePr.summaryBelow is False
    analytics_levels = [int(analytics_sheet.row_dimensions[index].outline_level or 0) for index in range(2, analytics_sheet.max_row + 1)]
    assert 0 in analytics_levels
    assert 1 in analytics_levels

    item_ids_export = client.get(
        "/approval-register/export",
        params={"item_ids": tagged.json()["id"]},
        headers=employee,
    )
    assert item_ids_export.status_code == 200, item_ids_export.text
    filtered = load_workbook(io.BytesIO(item_ids_export.content))
    filtered_names = [row[4].value for row in filtered["Строки"].iter_rows(min_row=2)]
    assert filtered_names == ["Tagged line"]


def test_register_export_uses_excel_outline_and_packs_files(tmp_path):
    client = make_client(tmp_path)
    employee = auth(client, "employee", "employee")
    request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    item = client.post(
        f"/requests/{request['id']}/items",
        json={"dds_id": DDS_LICENSE_ID, "name": "Line with file", "sum_plan": 100, "analytics_1": "Проект А"},
        headers=employee,
    ).json()
    uploaded = client.post(
        f"/items/{item['id']}/files",
        files={"file": ("evidence.png", b"test", "image/png")},
        headers=employee,
    )
    assert uploaded.status_code == 200, uploaded.text

    exported = client.get(
        "/approval-register/export",
        params={"request_id": request["id"], "view": "cfo"},
        headers=employee,
    )
    assert exported.status_code == 200, exported.text
    assert "zip" in exported.headers.get("content-type", "")

    with zipfile.ZipFile(io.BytesIO(exported.content)) as archive:
        names = archive.namelist()
        assert any(name.endswith(".xlsx") for name in names)
        assert any(name.endswith("evidence.png") for name in names)
        assert any(name.startswith("Приложения/") for name in names)
        xlsx_name = next(name for name in names if name.endswith(".xlsx"))
        workbook = load_workbook(io.BytesIO(archive.read(xlsx_name)))

    registry = workbook["Реестр"]
    assert registry.sheet_properties.outlinePr.summaryBelow is False
    outline_levels = [int(registry.row_dimensions[index].outline_level or 0) for index in range(2, registry.max_row + 1)]
    assert max(outline_levels) >= 1
    structure_names = [row[0].value for row in registry.iter_rows(min_row=2)]
    assert "Line with file" in structure_names
    headers = [cell.value for cell in registry[1]]
    assert "Файлы" in headers
    files_index = headers.index("Файлы")
    assert any(
        row[files_index].value and "evidence.png" in str(row[files_index].value)
        for row in registry.iter_rows(min_row=2)
    )

    details_headers = [cell.value for cell in workbook["Строки"][1]]
    assert "Приложение 1" in details_headers
    attachment_index = details_headers.index("Приложение 1")
    attachment_cell = next(
        row[attachment_index]
        for row in workbook["Строки"].iter_rows(min_row=2)
        if row[4].value == "Line with file"
    )
    assert attachment_cell.value == "evidence.png"
    assert attachment_cell.hyperlink is not None
    hyperlink_target = str(getattr(attachment_cell.hyperlink, "target", None) or attachment_cell.hyperlink)
    assert "Приложения/" in hyperlink_target

    summary_values = {
        row[0].value: row[1].value
        for row in workbook["Сводка"].iter_rows(min_row=2, max_col=2)
        if row[0].value
    }
    assert summary_values["Файлов"] == 1
