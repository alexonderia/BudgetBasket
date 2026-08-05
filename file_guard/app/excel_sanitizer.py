from __future__ import annotations

import io
import re
import zipfile
from copy import copy
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook

from .config import settings

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
FORMULA_VALUE_UNAVAILABLE_TEXT = "Значение формулы недоступно"
_FORMULA_STRING_LITERAL = re.compile(r'"((?:[^"]|"")*)"')


class ExcelSanitizationError(ValueError):
    """The workbook cannot be reduced to a safe viewing copy."""


@dataclass(frozen=True, slots=True)
class SanitizedWorkbook:
    content: bytes
    removed_components: tuple[str, ...]
    warnings: tuple[str, ...]


def sanitize_excel(content: bytes) -> SanitizedWorkbook:
    """Rebuild an OOXML workbook, deliberately copying only passive view data."""
    _check_encryption(content)
    removed = _detect_removed_components(content)
    source = None
    cached = None
    try:
        source = load_workbook(io.BytesIO(content), read_only=False, data_only=False, keep_links=False)
        cached = load_workbook(io.BytesIO(content), read_only=False, data_only=True, keep_links=False)
    except Exception as exc:
        raise ExcelSanitizationError("Workbook cannot be safely read") from exc

    if not source.worksheets or len(source.worksheets) > settings.excel_max_sheets:
        raise ExcelSanitizationError("Workbook sheet limit exceeded")

    result = Workbook()
    result.remove(result.active)
    warnings: set[str] = set()
    non_empty_cells = 0

    try:
        for source_sheet, cached_sheet in zip(source.worksheets, cached.worksheets, strict=True):
            if source_sheet.max_row > settings.excel_max_rows_per_sheet or source_sheet.max_column > settings.excel_max_columns_per_sheet:
                raise ExcelSanitizationError("Workbook dimensions exceed limit")
            if len(source_sheet.merged_cells.ranges) > settings.excel_max_merged_ranges:
                raise ExcelSanitizationError("Workbook merge limit exceeded")

            target = result.create_sheet(title=source_sheet.title)
            target.sheet_state = source_sheet.sheet_state
            target.freeze_panes = source_sheet.freeze_panes
            for key, dimension in source_sheet.column_dimensions.items():
                target.column_dimensions[key].width = dimension.width
                target.column_dimensions[key].hidden = dimension.hidden
            for row_number, dimension in source_sheet.row_dimensions.items():
                target.row_dimensions[row_number].height = dimension.height
                target.row_dimensions[row_number].hidden = dimension.hidden

            for row in source_sheet.iter_rows():
                for source_cell in row:
                    value = source_cell.value
                    if value is None:
                        continue
                    non_empty_cells += 1
                    if non_empty_cells > settings.excel_max_non_empty_cells:
                        raise ExcelSanitizationError("Workbook cell limit exceeded")
                    target_cell = target.cell(row=source_cell.row, column=source_cell.column)
                    if source_cell.data_type == "f":
                        value = cached_sheet.cell(source_cell.row, source_cell.column).value
                        if value is None:
                            # Preserve the human-facing label where one exists
                            # (for example, HYPERLINK(url, "label")), but never
                            # copy the executable expression or its URL/path.
                            value = _formula_text_value(source_cell.value)
                            warnings.add("FORMULA_WITHOUT_CACHED_VALUE")
                        _set_as_text_if_formula(target_cell, value)
                    else:
                        _set_as_text_if_formula(target_cell, value)
                    if isinstance(value, str) and len(value) > settings.excel_max_cell_string_length:
                        raise ExcelSanitizationError("Workbook cell string is too large")
                    target_cell.font = copy(source_cell.font)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.border = copy(source_cell.border)
                    target_cell.alignment = copy(source_cell.alignment)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = copy(source_cell.protection)
            for merged_range in source_sheet.merged_cells.ranges:
                target.merge_cells(str(merged_range))

        # Excel forbids a workbook in which every sheet is hidden.
        if all(sheet.sheet_state != "visible" for sheet in result.worksheets):
            result.worksheets[0].sheet_state = "visible"
            warnings.add("ALL_SHEETS_HIDDEN")
        buffer = io.BytesIO()
        result.save(buffer)
    except ExcelSanitizationError:
        raise
    except Exception as exc:
        raise ExcelSanitizationError("Workbook cannot be safely rebuilt") from exc
    finally:
        if source is not None:
            source.close()
        if cached is not None:
            cached.close()

    output = buffer.getvalue()
    if len(output) > settings.excel_max_output_size_bytes:
        raise ExcelSanitizationError("Sanitized workbook exceeds size limit")
    verify_sanitized_excel(output)
    return SanitizedWorkbook(output, tuple(sorted(removed)), tuple(sorted(warnings)))


def verify_sanitized_excel(content: bytes) -> None:
    """Confirm the generated archive has no active OOXML parts or formulas."""
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            names = {name.lower() for name in archive.namelist()}
            required = {"[content_types].xml", "xl/workbook.xml"}
            prohibited = ("vbaproject", "activex", "embeddings/", "externallinks/", "connections.xml", "oleobject")
            if not required.issubset(names) or any(part in name for name in names for part in prohibited):
                raise ExcelSanitizationError("Sanitized archive contains prohibited parts")
            for name in archive.namelist():
                if name.lower().endswith(".rels") and b'TargetMode="External"' in archive.read(name):
                    raise ExcelSanitizationError("Sanitized archive contains external relationship")
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False, keep_links=False)
        try:
            if any(cell.data_type == "f" for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row):
                raise ExcelSanitizationError("Sanitized workbook contains formulas")
        finally:
            workbook.close()
    except ExcelSanitizationError:
        raise
    except Exception as exc:
        raise ExcelSanitizationError("Sanitized workbook is invalid") from exc


def _set_as_text_if_formula(cell, value) -> None:
    cell.value = value
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"


def _formula_text_value(formula: object) -> str:
    if not isinstance(formula, str):
        return FORMULA_VALUE_UNAVAILABLE_TEXT
    literals = [match.replace('""', '"') for match in _FORMULA_STRING_LITERAL.findall(formula)]
    # The second HYPERLINK argument is its display text; the last literal is a
    # safe and useful label for other formulas that contain a textual result.
    if literals:
        return literals[-1]
    return FORMULA_VALUE_UNAVAILABLE_TEXT


def _check_encryption(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            if any(info.flag_bits & 0x1 for info in archive.infolist()):
                raise ExcelSanitizationError("Encrypted workbook is not allowed")
    except ExcelSanitizationError:
        raise
    except zipfile.BadZipFile as exc:
        raise ExcelSanitizationError("Workbook is not a valid archive") from exc


def _detect_removed_components(content: bytes) -> set[str]:
    removed: set[str] = set()
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            names = [name.lower() for name in archive.namelist()]
            if any("vbaproject" in name for name in names): removed.add("vba_macros")
            if any("activex" in name for name in names): removed.add("active_x")
            if any("embeddings/" in name or "oleobject" in name for name in names): removed.add("ole")
            if any("externallinks/" in name or "connections" in name for name in names): removed.add("external_links")
            if any("media/" in name for name in names): removed.add("images")
            if any("charts/" in name for name in names): removed.add("charts")
    except zipfile.BadZipFile:
        pass
    return removed
